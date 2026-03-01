#!/usr/bin/env python3
"""
Rental Property Occupancy Tracker

Scrapes availability data from weekendica.com properties and stores it in Excel
and/or Supabase PostgreSQL. Run daily to build historical occupancy data.

Environment variables:
    STORAGE_MODE: "both" (Excel + DB, default), "db_only" (DB only, for CI)
    SUPABASE_URL: Supabase project URL
    SUPABASE_KEY: Supabase secret key (for write access)
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import logging
import time
import os
from pathlib import Path

# Configuration
EXCEL_FILE = Path(__file__).parent / "KuceZaIzdavanje.xlsx"
PROPERTIES_SHEET = "Properties"
AVAILABILITY_SHEET = "Availability"
REQUEST_DELAY = 1  # seconds between requests to be polite to the server
STORAGE_MODE = os.environ.get("STORAGE_MODE", "both")  # "both" or "db_only"

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def setup_excel_structure():
    """Ensure Excel file has correct structure with ID column and Availability sheet."""
    logger.info(f"Checking Excel structure: {EXCEL_FILE}")

    # Load workbook
    wb = load_workbook(EXCEL_FILE)

    # Get first sheet (properties data)
    first_sheet = wb.active
    first_sheet.title = PROPERTIES_SHEET

    # Check if ID column exists (should be column A)
    has_id_column = first_sheet.cell(row=1, column=1).value == "ID"

    if not has_id_column:
        logger.info("Adding ID column to Properties sheet")
        # Insert new column A for ID
        first_sheet.insert_cols(1)
        first_sheet.cell(row=1, column=1, value="ID")

        # Add IDs (1-indexed, starting from row 2)
        for row_num in range(2, first_sheet.max_row + 1):
            first_sheet.cell(row=row_num, column=1, value=row_num - 1)

    # Create Availability sheet if it doesn't exist
    if AVAILABILITY_SHEET not in wb.sheetnames:
        logger.info("Creating Availability sheet")
        avail_sheet = wb.create_sheet(AVAILABILITY_SHEET)
        # Add headers
        headers = ["property_id", "date", "booked", "checked_at", "day_of_week", "month_of_year"]
        for col, header in enumerate(headers, 1):
            avail_sheet.cell(row=1, column=col, value=header)

    wb.save(EXCEL_FILE)
    logger.info("Excel structure verified/updated")


def get_properties() -> list[dict]:
    """Read properties from Excel and return list with ID and URL."""
    df = pd.read_excel(EXCEL_FILE, sheet_name=PROPERTIES_SHEET)

    # URL column is "Vikendice" (second column after ID)
    url_col = "Vikendice" if "Vikendice" in df.columns else df.columns[1]

    properties = []
    for _, row in df.iterrows():
        # Skip rows with missing ID or URL
        if pd.isna(row["ID"]) or pd.isna(row[url_col]):
            continue

        url = str(row[url_col])

        # Only include valid weekendica.com URLs
        if not url.startswith("https://www.weekendica.com/"):
            continue

        properties.append({
            "id": int(row["ID"]),
            "url": url
        })

    return properties


def fetch_calendar_data(session: requests.Session, url: str) -> dict[str, bool]:
    """
    Fetch property page and parse calendar availability.

    Returns dict: {date_str: is_booked} e.g., {"2026-01-19": True, "2026-01-20": False}
    """
    logger.info(f"Fetching: {url}")

    try:
        response = session.get(url, timeout=30)
        response.raise_for_status()
    except requests.RequestException as e:
        logger.error(f"Failed to fetch {url}: {e}")
        return {}

    soup = BeautifulSoup(response.text, "html.parser")

    # Find all calendar day elements
    # Available: class contains "rz--available"
    # Booked: class contains "rz--not-available" or "rz--day-unavailable"
    calendar_days = soup.find_all("li", attrs={"data-date": True})

    availability = {}
    for day in calendar_days:
        date_str = day.get("data-date")  # Format: "29-04-2026"
        if not date_str:
            continue

        # Convert from DD-MM-YYYY to YYYY-MM-DD
        try:
            date_obj = datetime.strptime(date_str, "%d-%m-%Y")
            date_iso = date_obj.strftime("%Y-%m-%d")
        except ValueError:
            continue

        # Check if booked (not available)
        classes = day.get("class", [])
        is_booked = "rz--available" not in classes

        availability[date_iso] = is_booked

    return availability


def get_target_dates() -> list[str]:
    """Get today and tomorrow's dates in YYYY-MM-DD format."""
    today = datetime.now()
    tomorrow = today + timedelta(days=1)
    return [
        today.strftime("%Y-%m-%d"),
        tomorrow.strftime("%Y-%m-%d")
    ]


def batch_update_availability(updates: list[tuple[int, str, bool]]):
    """Batch update or insert availability records in Excel.

    Opens the workbook once, builds an index for O(1) lookups, and saves once.

    Args:
        updates: List of (property_id, date_str, booked) tuples.
    """
    wb = load_workbook(EXCEL_FILE)
    sheet = wb[AVAILABILITY_SHEET]
    checked_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Build index: (property_id, date_str) -> row_number (single pass)
    existing_rows = {}
    for row_num in range(2, sheet.max_row + 1):
        key = (sheet.cell(row=row_num, column=1).value,
               str(sheet.cell(row=row_num, column=2).value)[:10])
        existing_rows[key] = row_num

    updated = 0
    inserted = 0
    for property_id, date, booked in updates:
        date_obj = datetime.strptime(date, "%Y-%m-%d")
        day_of_week = date_obj.strftime("%A")
        month_of_year = date_obj.strftime("%B")

        key = (property_id, date)
        row = existing_rows.get(key)

        if row:
            sheet.cell(row=row, column=3, value=1 if booked else 0)
            sheet.cell(row=row, column=4, value=checked_at)
            sheet.cell(row=row, column=5, value=day_of_week)
            sheet.cell(row=row, column=6, value=month_of_year)
            updated += 1
        else:
            new_row = sheet.max_row + 1
            sheet.cell(row=new_row, column=1, value=property_id)
            sheet.cell(row=new_row, column=2, value=date)
            sheet.cell(row=new_row, column=3, value=1 if booked else 0)
            sheet.cell(row=new_row, column=4, value=checked_at)
            sheet.cell(row=new_row, column=5, value=day_of_week)
            sheet.cell(row=new_row, column=6, value=month_of_year)
            existing_rows[key] = new_row
            inserted += 1

    wb.save(EXCEL_FILE)
    logger.info(f"Availability batch update: {updated} updated, {inserted} inserted")


def init_supabase():
    """Initialize Supabase client from environment variables."""
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_KEY")
    if not url or not key:
        logger.warning("SUPABASE_URL or SUPABASE_KEY not set, database storage disabled")
        return None
    from supabase import create_client
    return create_client(url, key)


def get_properties_from_db(db) -> list[dict]:
    """Read properties from Supabase and return list with ID and URL."""
    result = db.table("properties").select("id, url").order("id").execute()
    return [{"id": r["id"], "url": r["url"]} for r in result.data]


def batch_upsert_supabase(db, updates: list[tuple[int, str, bool]]):
    """Upsert availability records to Supabase."""
    checked_at = datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    records = [
        {"property_id": prop_id, "date": date, "booked": booked, "checked_at": checked_at}
        for prop_id, date, booked in updates
    ]
    result = db.table("availability").upsert(
        records, on_conflict="property_id,date"
    ).execute()
    logger.info(f"Supabase upsert: {len(result.data)} rows")
    if len(result.data) != len(records):
        logger.warning(f"Supabase upsert mismatch: sent {len(records)}, got {len(result.data)} — possible RLS issue")


def calculate_occupancy_summaries():
    """Calculate monthly occupancy percentages and update Properties sheet."""
    logger.info("Calculating occupancy summaries")

    # Read availability data
    try:
        avail_df = pd.read_excel(EXCEL_FILE, sheet_name=AVAILABILITY_SHEET)
    except (ValueError, KeyError):
        logger.warning("No availability data yet, skipping summary calculation")
        return

    if avail_df.empty:
        return

    # Convert date column to datetime
    avail_df["date"] = pd.to_datetime(avail_df["date"])
    avail_df["year_month"] = avail_df["date"].dt.to_period("M")

    # Calculate occupancy per property per month
    monthly_occ = avail_df.groupby(["property_id", "year_month"])["booked"].mean()

    # Calculate weekend vs weekday occupancy
    weekend_days = ["Friday", "Saturday", "Sunday"]
    weekday_days = ["Monday", "Tuesday", "Wednesday", "Thursday"]

    weekend_df = avail_df[avail_df["day_of_week"].isin(weekend_days)]
    weekday_df = avail_df[avail_df["day_of_week"].isin(weekday_days)]

    weekend_occ = weekend_df.groupby("property_id")["booked"].mean() if not weekend_df.empty else pd.Series(dtype=float)
    weekday_occ = weekday_df.groupby("property_id")["booked"].mean() if not weekday_df.empty else pd.Series(dtype=float)

    # Calculate total occupancy per property
    total_occ = avail_df.groupby("property_id")["booked"].mean()

    # Calculate rankings (1 = highest occupancy)
    rankings = total_occ.rank(ascending=False, method="min").astype(int)

    # Load workbook to update Properties sheet
    wb = load_workbook(EXCEL_FILE)
    props_sheet = wb[PROPERTIES_SHEET]

    # Find existing occupancy columns or add new ones
    header_row = 1
    existing_headers = [props_sheet.cell(row=header_row, column=col).value
                       for col in range(1, props_sheet.max_column + 1)]

    # Helper function to get or create column
    def get_or_create_col(col_name):
        if col_name in existing_headers:
            return existing_headers.index(col_name) + 1
        col_idx = props_sheet.max_column + 1
        props_sheet.cell(row=header_row, column=col_idx, value=col_name)
        existing_headers.append(col_name)
        return col_idx

    # Helper to write a data series into a named column
    def write_column(col_name, data_series, formatter=lambda v: f"{v:.0%}", fallback="0%"):
        col_idx = get_or_create_col(col_name)
        for row_num in range(2, props_sheet.max_row + 1):
            prop_id = props_sheet.cell(row=row_num, column=1).value
            if prop_id is None:
                continue
            value = data_series.get(prop_id, None)
            props_sheet.cell(row=row_num, column=col_idx,
                             value=formatter(value) if value is not None else fallback)

    # Summary columns
    write_column("Occ_Weekend", weekend_occ)
    write_column("Occ_Weekday", weekday_occ)
    write_column("Occ_Total", total_occ)
    write_column("Rank", rankings, formatter=lambda v: int(v), fallback="-")

    # Monthly occupancy columns (stack as new months are added)
    for month in sorted(avail_df["year_month"].unique()):
        month_data = monthly_occ.xs(month, level="year_month", drop_level=True)
        write_column(f"Occ_{month}", month_data)

    wb.save(EXCEL_FILE)
    logger.info("Occupancy summaries updated")


def main():
    """Main entry point for the scraper."""
    logger.info("=" * 50)
    logger.info(f"Starting occupancy scraper (storage: {STORAGE_MODE})")
    logger.info("=" * 50)

    use_excel = STORAGE_MODE in ("both", "excel_only")
    use_db = STORAGE_MODE in ("both", "db_only")

    # Initialize Supabase client if needed
    db = init_supabase() if use_db else None
    if use_db and not db:
        logger.error("Database storage requested but Supabase credentials not set")
        return

    # Ensure Excel structure is correct (skip in db_only mode)
    if use_excel:
        setup_excel_structure()

    # Get list of properties
    if use_excel:
        properties = get_properties()
    else:
        properties = get_properties_from_db(db)
    logger.info(f"Found {len(properties)} properties")

    # Get target dates (today and tomorrow)
    target_dates = get_target_dates()
    logger.info(f"Target dates: {target_dates}")

    # Scrape each property
    session = requests.Session()
    session.headers["User-Agent"] = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"

    success_count = 0
    availability_updates = []
    for prop in properties:
        prop_id = prop["id"]
        url = prop["url"]

        # Fetch calendar data
        calendar_data = fetch_calendar_data(session, url)

        # Always delay between requests, even on failure
        time.sleep(REQUEST_DELAY)

        if not calendar_data:
            logger.warning(f"No calendar data for property {prop_id}")
            continue

        # Collect availability for target dates
        for date in target_dates:
            if date in calendar_data:
                availability_updates.append((prop_id, date, calendar_data[date]))
            else:
                logger.warning(f"Date {date} not found in calendar for property {prop_id}")

        success_count += 1

    logger.info(f"Successfully scraped {success_count}/{len(properties)} properties")

    if availability_updates:
        # Write to Excel
        if use_excel:
            batch_update_availability(availability_updates)

        # Write to Supabase
        if use_db and db:
            try:
                batch_upsert_supabase(db, availability_updates)
            except Exception as e:
                logger.error(f"Supabase write failed: {e}")

    # Calculate and update occupancy summaries (Excel only)
    if use_excel:
        calculate_occupancy_summaries()

    logger.info("Scraper finished")


if __name__ == "__main__":
    main()
